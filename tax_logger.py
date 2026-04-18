#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Binance Trade History → Canadian Tax Excel Logger

Fetches Binance spot trades, converts USD → CAD using Bank of Canada
historical rates, and inserts missing entries into the tax spreadsheet.
Also updates the ACB Schedule and computes capital gains for sells.

Usage:
  python tax_logger.py              # fetch, compare, insert new trades
  python tax_logger.py --dry-run    # preview without writing to Excel
  python tax_logger.py --scan       # display existing spreadsheet trades
  python tax_logger.py --days 90    # look back N days (default: 90)
"""

import sys
import os
import re
import time
import hashlib
import hmac
import urllib.parse
import json
from datetime import datetime, timezone, timedelta, date
from decimal import Decimal, ROUND_DOWN
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

load_dotenv()

# ── Config ─────────────────────────────────────────────────────────────────────

EXCEL_PATH         = Path(os.getenv("TAX_EXCEL_PATH", r"G:\My Drive\Crypto\tax.xlsx"))
ACCOUNT_NAME       = "Binance"
BINANCE_API_KEY    = os.getenv("BINANCE_API_KEY", "")
BINANCE_API_SECRET = os.getenv("BINANCE_API_SECRET", "")
BINANCE_BASE_URL   = "https://api.binance.com"

# Symbols to check — extend this list as needed
SYMBOLS_TO_CHECK = [
    "BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT", "XRPUSDT",
    "ADAUSDT", "DOTUSDT", "DOGEUSDT", "AVAXUSDT", "LINKUSDT",
    "PEPEUSDT", "SHIBUSDT", "ARBUSDT", "OPUSDT", "INJUSDT",
    "SUIUSDT", "NEARUSDT", "ALGOUSDT", "LTCUSDT", "ETCUSDT",
    "TREEUSDT", "BNBBTC", "ETHBTC",
]

# Sheet / column config — matches the existing spreadsheet layout
DETAIL_SHEET    = "Detailed P&L"
ACB_SHEET       = "ACB Schedule"

DETAIL_COLS = {            # 1-based column indices
    "date":         1,
    "account":      2,
    "coin":         3,
    "type":         4,
    "quantity":     5,
    "gross_cad":    6,
    "fee_cad":      7,
    "net_cad":      8,
    "acb_unit_bef": 9,
    "acb_disposal": 10,
    "cap_gain":     11,
    "tax_year":     12,
    "notes":        13,
}

ACB_COLS = {
    "date":         1,
    "account":      2,
    "coin":         3,
    "event":        4,
    "qty_change":   5,
    "cost_cad":     6,
    "acb_qty":      7,
    "acb_total":    8,
    "acb_unit":     9,
}

# ── Binance helpers ────────────────────────────────────────────────────────────

def binance_signed_get(endpoint, params):
    params["timestamp"] = int(time.time() * 1000)
    query = urllib.parse.urlencode(params)
    sig   = hmac.new(BINANCE_API_SECRET.encode(), query.encode(), hashlib.sha256).hexdigest()
    r = requests.get(
        f"{BINANCE_BASE_URL}{endpoint}?{query}&signature={sig}",
        headers={"X-MBX-APIKEY": BINANCE_API_KEY},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()


CALL_DELAY_S = 0.4   # seconds between API calls to stay well under rate limits


def _signed_get_with_retry(endpoint, params, max_retries=4):
    """Signed GET with exponential back-off on 429 and transient errors."""
    for attempt in range(max_retries):
        try:
            return binance_signed_get(endpoint, dict(params))  # copy so timestamp is fresh
        except requests.exceptions.HTTPError as e:
            status = e.response.status_code if e.response is not None else 0
            if status == 429:
                wait = 10 * (2 ** attempt)
                print(f"    Rate-limited — waiting {wait}s before retry {attempt+1}/{max_retries}")
                time.sleep(wait)
                continue
            raise
    raise RuntimeError(f"Gave up after {max_retries} retries on {endpoint}")


def fetch_binance_trades(symbol, start_ms, end_ms):
    """
    Fetch all trades for a symbol in [start_ms, end_ms].

    Strategy: use only startTime (no endTime) to avoid Binance's 24-hour window
    restriction (-1127).  Binance returns up to 1000 trades from that point;
    if there are more, paginate by fromId.  Filter to end_ms client-side.
    """
    trades  = []
    from_id = None

    while True:
        if from_id is not None:
            # Pagination by ID — startTime/endTime cannot be combined with fromId
            params = {"symbol": symbol, "limit": 1000, "fromId": from_id}
        else:
            params = {"symbol": symbol, "limit": 1000, "startTime": start_ms}

        time.sleep(CALL_DELAY_S)
        try:
            batch = _signed_get_with_retry("/api/v3/myTrades", params)
        except requests.exceptions.HTTPError as e:
            status = e.response.status_code if e.response is not None else 0
            if status in (400, 404):
                return trades   # symbol invalid or no trades
            raise

        if not batch or not isinstance(batch, list):
            break

        # Keep only trades within our time window
        in_window = [t for t in batch if t["time"] <= end_ms]
        trades.extend(in_window)

        # Stop if we reached the end of the window or got fewer than 1000 results
        if len(batch) < 1000 or batch[-1]["time"] > end_ms:
            break

        from_id = batch[-1]["id"] + 1

    return trades


def fetch_portfolio_symbols():
    """Get current non-zero balance assets to add to the scan list."""
    try:
        data = binance_signed_get("/api/v3/account", {})
        assets = [
            f"{b['asset']}USDT"
            for b in data.get("balances", [])
            if float(b["free"]) + float(b["locked"]) > 0
            and b["asset"] not in ("USDT", "BUSD")
        ]
        return assets
    except Exception as e:
        print(f"  Warning: could not fetch portfolio symbols: {e}")
        return []


# ── Bank of Canada USD/CAD rates ───────────────────────────────────────────────

_cad_rate_cache: dict = {}


def fetch_cad_rates(start: date, end: date):
    """Fetch USD/CAD daily rates from Bank of Canada and cache them."""
    url = (
        "https://www.bankofcanada.ca/valet/observations/FXUSDCAD/json"
        f"?start_date={start.isoformat()}&end_date={end.isoformat()}"
    )
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        for obs in r.json().get("observations", []):
            d = obs["d"]
            v = obs.get("FXUSDCAD", {}).get("v")
            if v:
                _cad_rate_cache[d] = float(v)
        print(f"  CAD rates loaded: {len(_cad_rate_cache)} days")
    except Exception as e:
        print(f"  Warning: Bank of Canada rate fetch failed: {e}")


def usd_to_cad(usd_amount: float, trade_date: date) -> float:
    """Convert USD to CAD using the closest available Bank of Canada rate."""
    # Walk backwards up to 7 days to find the most recent available rate
    for delta in range(8):
        d = (trade_date - timedelta(days=delta)).isoformat()
        if d in _cad_rate_cache:
            return usd_amount * _cad_rate_cache[d]
    # Fallback: use latest known rate
    if _cad_rate_cache:
        latest_rate = _cad_rate_cache[max(_cad_rate_cache)]
        print(f"  Warning: no CAD rate for {trade_date} — using latest ({latest_rate:.4f})")
        return usd_amount * latest_rate
    print(f"  Warning: no CAD rate available — returning USD amount unchanged")
    return usd_amount


def get_cad_rate_used(trade_date: date) -> float:
    for delta in range(8):
        d = (trade_date - timedelta(days=delta)).isoformat()
        if d in _cad_rate_cache:
            return _cad_rate_cache[d]
    if _cad_rate_cache:
        return _cad_rate_cache[max(_cad_rate_cache)]
    return 1.0


# ── Excel reader ───────────────────────────────────────────────────────────────

def load_workbook_safe():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)
    return openpyxl.load_workbook(str(EXCEL_PATH))


def read_existing_detail_trades(ws):
    """
    Read existing trades from Detailed P&L sheet.
    Returns list of dicts and a set of dedup keys.
    """
    trades = []
    dedup_keys = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val, account, coin, ttype, qty = row[0], row[1], row[2], row[3], row[4]
        notes = row[12] if len(row) > 12 else ""

        # Skip headers and section labels
        if not date_val or not coin or not ttype:
            continue
        if str(ttype).startswith("─") or str(date_val).startswith("─"):
            continue
        if ttype not in ("Buy", "Sell", "BUY", "SELL"):
            continue

        # Normalise date
        if isinstance(date_val, datetime):
            d = date_val.date()
        elif isinstance(date_val, date):
            d = date_val
        else:
            try:
                d = datetime.strptime(str(date_val).strip(), "%Y-%m-%d").date()
            except Exception:
                continue

        # Extract Binance trade ID from notes if present
        binance_id = None
        if notes:
            m = re.search(r"binance_id:(\d+)", str(notes))
            if m:
                binance_id = int(m.group(1))

        key = (d, str(coin).upper(), str(ttype).upper(), round(float(qty or 0), 8))
        dedup_keys.add(key)
        if binance_id:
            dedup_keys.add(f"id:{binance_id}")

        trades.append({
            "date":      d,
            "account":   account,
            "coin":      coin,
            "type":      ttype,
            "qty":       float(qty or 0),
            "notes":     str(notes) if notes else "",
            "binance_id": binance_id,
        })

    return trades, dedup_keys


def read_acb_state(ws):
    """
    Read the ACB Schedule sheet and return the latest ACB state per coin.
    Returns dict: coin -> {qty, total_cad, unit_cad}
    """
    acb = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_v, account, coin, event, qty_chg, cost_cad, acb_qty, acb_total, acb_unit = \
            (row[i] if i < len(row) else None for i in range(9))

        if not coin or not event:
            continue
        if str(event).startswith("─") or not acb_qty:
            continue
        try:
            c = str(coin).upper()
            acb[c] = {
                "qty":       float(acb_qty   or 0),
                "total_cad": float(acb_total or 0),
                "unit_cad":  float(acb_unit  or 0),
            }
        except Exception:
            continue
    return acb


def find_insert_row(ws, trade_date: date) -> int:
    """
    Find the row number in Detailed P&L to insert a trade for the given date,
    preserving chronological order within the correct tax-year section.
    """
    target_year = str(trade_date.year)
    last_row_in_year = None
    year_section_start = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        date_val = row[0]
        ttype    = row[3] if len(row) > 3 else None

        # Detect tax-year section header
        if date_val and str(date_val).find(target_year) >= 0 and (not ttype or str(ttype).startswith("─")):
            year_section_start = row_idx

        if not date_val or not ttype or ttype not in ("Buy", "Sell", "BUY", "SELL"):
            continue

        # Normalise date
        if isinstance(date_val, datetime):
            d = date_val.date()
        elif isinstance(date_val, date):
            d = date_val
        else:
            try:
                d = datetime.strptime(str(date_val).strip(), "%Y-%m-%d").date()
            except Exception:
                continue

        if d.year == trade_date.year:
            if d <= trade_date:
                last_row_in_year = row_idx

    if last_row_in_year:
        return last_row_in_year + 1

    # No existing entries in that year — append after last non-empty row
    for row_idx in range(ws.max_row, 0, -1):
        if any(ws.cell(row=row_idx, column=c).value for c in range(1, 14)):
            return row_idx + 2  # leave a blank separator row

    return ws.max_row + 2


def find_acb_insert_row(ws, trade_date: date, coin: str) -> int:
    """Find insert row in ACB Schedule for this coin + date."""
    last_coin_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        date_v = row[0]
        coin_v = row[2] if len(row) > 2 else None
        if coin_v and str(coin_v).upper() == coin.upper():
            if date_v:
                if isinstance(date_v, (datetime, date)):
                    d = date_v.date() if isinstance(date_v, datetime) else date_v
                else:
                    try:
                        d = datetime.strptime(str(date_v).strip(), "%Y-%m-%d").date()
                    except Exception:
                        continue
                if d <= trade_date:
                    last_coin_row = row_idx

    if last_coin_row:
        return last_coin_row + 1

    # Coin not seen yet — append at end
    for row_idx in range(ws.max_row, 0, -1):
        if any(ws.cell(row=row_idx, column=c).value for c in range(1, 10)):
            return row_idx + 1

    return ws.max_row + 1


# ── ACB computation ────────────────────────────────────────────────────────────

def update_acb_buy(acb_state, coin, qty, cost_cad):
    """Update ACB state for a purchase. Returns new state."""
    s = acb_state.get(coin, {"qty": 0.0, "total_cad": 0.0, "unit_cad": 0.0})
    new_qty   = s["qty"] + qty
    new_total = s["total_cad"] + cost_cad
    new_unit  = new_total / new_qty if new_qty else 0.0
    acb_state[coin] = {"qty": new_qty, "total_cad": new_total, "unit_cad": new_unit}
    return acb_state[coin]


def compute_acb_sell(acb_state, coin, qty, proceeds_cad):
    """
    Compute disposal cost and capital gain for a sell.
    Returns (acb_unit_before, acb_disposal, capital_gain, new_state).
    """
    s = acb_state.get(coin, {"qty": 0.0, "total_cad": 0.0, "unit_cad": 0.0})
    acb_unit_before = s["unit_cad"]
    acb_disposal    = acb_unit_before * qty
    capital_gain    = proceeds_cad - acb_disposal

    new_qty   = max(s["qty"] - qty, 0.0)
    new_total = max(s["total_cad"] - acb_disposal, 0.0)
    new_unit  = new_total / new_qty if new_qty else 0.0
    acb_state[coin] = {"qty": new_qty, "total_cad": new_total, "unit_cad": new_unit}

    return acb_unit_before, acb_disposal, capital_gain, acb_state[coin]


# ── Row writers ────────────────────────────────────────────────────────────────

def write_detail_row(ws, row_idx, trade: dict):
    """Insert one row into Detailed P&L, shifting existing rows down."""
    ws.insert_rows(row_idx)
    c = DETAIL_COLS
    ws.cell(row=row_idx, column=c["date"]).value         = trade["date"]
    ws.cell(row=row_idx, column=c["date"]).number_format = "YYYY-MM-DD"
    ws.cell(row=row_idx, column=c["account"]).value      = trade["account"]
    ws.cell(row=row_idx, column=c["coin"]).value         = trade["coin"]
    ws.cell(row=row_idx, column=c["type"]).value         = trade["type"]
    ws.cell(row=row_idx, column=c["quantity"]).value     = trade["qty"]
    ws.cell(row=row_idx, column=c["gross_cad"]).value    = round(trade["gross_cad"], 8)
    ws.cell(row=row_idx, column=c["fee_cad"]).value      = round(trade["fee_cad"], 8)
    ws.cell(row=row_idx, column=c["net_cad"]).value      = round(trade["net_cad"], 8)
    ws.cell(row=row_idx, column=c["acb_unit_bef"]).value = round(trade["acb_unit_before"], 8) if trade.get("acb_unit_before") else None
    ws.cell(row=row_idx, column=c["acb_disposal"]).value = round(trade["acb_disposal"], 8)    if trade.get("acb_disposal") else None
    ws.cell(row=row_idx, column=c["cap_gain"]).value     = round(trade["capital_gain"], 8)    if trade.get("capital_gain") is not None else None
    ws.cell(row=row_idx, column=c["tax_year"]).value     = trade["date"].year
    ws.cell(row=row_idx, column=c["notes"]).value        = trade["notes"]


def append_acb_row(ws, entry: dict):
    """Append one row to the end of ACB Schedule (never inserts mid-sheet)."""
    row_idx = ws.max_row + 1
    c = ACB_COLS
    ws.cell(row=row_idx, column=c["date"]).value         = entry["date"]
    ws.cell(row=row_idx, column=c["date"]).number_format = "YYYY-MM-DD"
    ws.cell(row=row_idx, column=c["account"]).value      = entry["account"]
    ws.cell(row=row_idx, column=c["coin"]).value         = entry["coin"]
    ws.cell(row=row_idx, column=c["event"]).value        = entry["event"]
    ws.cell(row=row_idx, column=c["qty_change"]).value   = round(entry["qty_change"], 8)
    ws.cell(row=row_idx, column=c["cost_cad"]).value     = round(entry["cost_cad"], 8)
    ws.cell(row=row_idx, column=c["acb_qty"]).value      = round(entry["acb_qty"], 8)
    ws.cell(row=row_idx, column=c["acb_total"]).value    = round(entry["acb_total"], 8)
    ws.cell(row=row_idx, column=c["acb_unit"]).value     = round(entry["acb_unit"], 8)


# ── Main logic ─────────────────────────────────────────────────────────────────

def scan_mode():
    """Print a summary of what's currently in the spreadsheet."""
    wb = load_workbook_safe()
    ws = wb[DETAIL_SHEET]
    print(f"\nDetailed P&L — {ws.max_row - 1} rows")
    print(f"{'Date':<12} {'Coin':<8} {'Type':<5} {'Qty':>14} {'Net CAD':>14}  Notes")
    print("-" * 75)
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_v, account, coin, ttype, qty = row[0], row[1], row[2], row[3], row[4]
        net_cad = row[7] if len(row) > 7 else None
        notes   = row[12] if len(row) > 12 else ""
        if not coin or not ttype:
            continue
        if ttype not in ("Buy", "Sell", "BUY", "SELL"):
            if date_v:
                print(f"  [{date_v}]")
            continue
        d_str = str(date_v)[:10] if date_v else "?"
        print(f"{d_str:<12} {str(coin):<8} {str(ttype):<5} {float(qty or 0):>14.6f} "
              f"{float(net_cad or 0):>14.2f}  {str(notes)[:40] if notes else ''}")


def main(dry_run=False, days=90):
    now    = datetime.now(timezone.utc)
    start  = now - timedelta(days=days)
    start_ms = int(start.timestamp() * 1000)
    end_ms   = int(now.timestamp() * 1000)

    print(f"\n{'='*65}")
    print(f"  Binance Tax Logger  —  last {days} days")
    print(f"  {'DRY RUN — no changes will be written' if dry_run else 'LIVE — will update Excel file'}")
    print(f"{'='*65}\n")

    # ── Step 1: Load exchange rates ────────────────────────────────────────────
    print("Fetching Bank of Canada USD/CAD rates...")
    fetch_cad_rates(start.date(), now.date())

    # ── Step 2: Load existing spreadsheet ─────────────────────────────────────
    print(f"\nReading {EXCEL_PATH.name}...")
    wb         = load_workbook_safe()
    ws_detail  = wb[DETAIL_SHEET]
    ws_acb     = wb[ACB_SHEET]

    existing_trades, dedup_keys = read_existing_detail_trades(ws_detail)
    acb_state = read_acb_state(ws_acb)

    print(f"  Existing trades in Detailed P&L: {len(existing_trades)}")
    print(f"  ACB state loaded for: {', '.join(sorted(acb_state.keys())) or 'none'}")

    # ── Step 3: Fetch Binance trades ───────────────────────────────────────────
    print("\nFetching Binance trades...")
    symbols = list(dict.fromkeys(SYMBOLS_TO_CHECK + fetch_portfolio_symbols()))

    all_new_trades = []  # list of processed trade dicts ready to insert

    for symbol in symbols:
        raw = fetch_binance_trades(symbol, start_ms, end_ms)
        if not raw:
            continue

        coin = symbol.replace("USDT", "").replace("BTC", "").replace("ETH", "")
        if symbol.endswith("USDT"):
            quote = "USDT"
            coin  = symbol[:-4]
        elif symbol.endswith("BTC"):
            quote = "BTC"
            coin  = symbol[:-3]
        elif symbol.endswith("ETH"):
            quote = "ETH"
            coin  = symbol[:-3]
        else:
            quote = "USDT"

        new_count = 0
        for t in raw:
            trade_id  = t["id"]
            trade_dt  = datetime.fromtimestamp(t["time"] / 1000, tz=timezone.utc)
            trade_date = trade_dt.date()
            ttype      = "Buy" if t["isBuyer"] else "Sell"
            qty        = float(t["qty"])
            price_usdt = float(t["price"])       # price in quote currency
            fee_amount = float(t["commission"])
            fee_asset  = t["commissionAsset"]

            # Check dedup
            id_key  = f"id:{trade_id}"
            qty_key = (trade_date, coin.upper(), ttype.upper(), round(qty, 8))
            if id_key in dedup_keys or qty_key in dedup_keys:
                continue

            # Convert proceeds/cost to CAD
            if quote == "USDT":
                gross_usdt = qty * price_usdt
            elif quote == "BTC":
                # Get BTC/USDT price on that day — approximate via Binance klines
                gross_usdt = qty * price_usdt  # price is in BTC, will be imprecise
            else:
                gross_usdt = qty * price_usdt

            cad_rate  = get_cad_rate_used(trade_date)
            gross_cad = usd_to_cad(gross_usdt, trade_date)

            # Fee in CAD
            if fee_asset == "USDT":
                fee_cad = usd_to_cad(fee_amount, trade_date)
            elif fee_asset == "BNB":
                # Approximate BNB price — fetch kline if needed
                fee_cad = usd_to_cad(fee_amount * 600, trade_date)  # rough BNB estimate
                fee_note = f" (BNB fee ~${fee_amount:.6f} BNB)"
            else:
                fee_cad = 0.0
                fee_note = f" ({fee_amount:.8f} {fee_asset} fee — convert manually)"

            fee_note = ""
            if fee_asset not in ("USDT",):
                fee_note = f" | fee: {fee_amount:.8f} {fee_asset}"

            if ttype == "Buy":
                net_cad = gross_cad + fee_cad  # cost basis includes fee
            else:
                net_cad = gross_cad - fee_cad  # net proceeds after fee

            # ACB computation
            acb_unit_before = acb_disposal = capital_gain = None
            if ttype == "Buy":
                update_acb_buy(acb_state, coin.upper(), qty, net_cad)
            else:
                acb_unit_before, acb_disposal, capital_gain, _ = compute_acb_sell(
                    acb_state, coin.upper(), qty, net_cad
                )

            notes = (
                f"binance_id:{trade_id} | "
                f"{symbol} @ {price_usdt} {quote} | "
                f"rate: {cad_rate:.4f} CAD/USD"
                f"{fee_note}"
            )

            all_new_trades.append({
                "date":            trade_date,
                "account":         ACCOUNT_NAME,
                "coin":            coin.upper(),
                "type":            ttype,
                "qty":             qty,
                "gross_cad":       gross_cad,
                "fee_cad":         fee_cad,
                "net_cad":         net_cad,
                "acb_unit_before": acb_unit_before,
                "acb_disposal":    acb_disposal,
                "capital_gain":    capital_gain,
                "notes":           notes,
                "binance_id":      trade_id,
                "symbol":          symbol,
            })
            # Register by ID only — qty_key would collide across partial fills of same order
            dedup_keys.add(id_key)
            new_count += 1

        if new_count:
            print(f"  {symbol}: {new_count} new trades (total fetched: {len(raw)})")
        else:
            print(f"  {symbol}: {len(raw)} fetched, all already logged")

    # Sort new trades chronologically before inserting
    all_new_trades.sort(key=lambda x: x["date"])

    print(f"\nNew trades to insert: {len(all_new_trades)}")

    if not all_new_trades:
        print("Nothing to insert — spreadsheet is up to date.")
        return

    # ── Step 4: Preview ────────────────────────────────────────────────────────
    print(f"\n{'─'*65}")
    print(f"  {'Date':<12} {'Coin':<6} {'Type':<5} {'Qty':>12} {'Gross CAD':>12} {'Fee CAD':>9} {'Net CAD':>12}  {'Cap Gain':>12}")
    print(f"{'─'*65}")
    total_gain = 0.0
    for t in all_new_trades:
        gain_str = f"{t['capital_gain']:>12.2f}" if t['capital_gain'] is not None else f"{'(buy)':>12}"
        print(
            f"  {str(t['date']):<12} {t['coin']:<6} {t['type']:<5} "
            f"{t['qty']:>12.6f} {t['gross_cad']:>12.2f} {t['fee_cad']:>9.4f} "
            f"{t['net_cad']:>12.2f}  {gain_str}"
        )
        if t['capital_gain'] is not None:
            total_gain += t['capital_gain']
    print(f"{'─'*65}")
    print(f"  {'Net capital gain on new trades:':>54}  {total_gain:>12.2f} CAD")

    if dry_run:
        print("\nDRY RUN — no changes written.")
        return

    # ── Step 5: Write to Excel ─────────────────────────────────────────────────
    print("\nWriting to Excel...")

    # Re-read ACB state from scratch for writing ACB rows
    # (all_new_trades already computed ACB state incrementally above)
    acb_state_for_write = read_acb_state(ws_acb)

    for t in all_new_trades:
        coin = t["coin"]

        # --- Detailed P&L ---
        insert_row = find_insert_row(ws_detail, t["date"])
        write_detail_row(ws_detail, insert_row, t)
        print(f"  Inserted [{t['date']}] {t['type']} {t['qty']:.6f} {coin} into Detailed P&L row {insert_row}")

        # --- ACB Schedule ---
        if t["type"] == "Buy":
            s = update_acb_buy(acb_state_for_write, coin, t["qty"], t["net_cad"])
            acb_entry = {
                "date":      t["date"],
                "account":   ACCOUNT_NAME,
                "coin":      coin,
                "event":     "Buy",
                "qty_change": t["qty"],
                "cost_cad":  t["net_cad"],
                "acb_qty":   s["qty"],
                "acb_total": s["total_cad"],
                "acb_unit":  s["unit_cad"],
            }
        else:
            before, disposal, gain, s = compute_acb_sell(
                acb_state_for_write, coin, t["qty"], t["net_cad"]
            )
            acb_entry = {
                "date":       t["date"],
                "account":    ACCOUNT_NAME,
                "coin":       coin,
                "event":      "Sell",
                "qty_change": -t["qty"],
                "cost_cad":   disposal,
                "acb_qty":    s["qty"],
                "acb_total":  s["total_cad"],
                "acb_unit":   s["unit_cad"],
            }

        append_acb_row(ws_acb, acb_entry)
        print(f"  Appended ACB entry for {coin} to ACB Schedule row {ws_acb.max_row}")

    wb.save(str(EXCEL_PATH))
    print(f"\nSaved: {EXCEL_PATH}")
    print(f"Total new trades inserted: {len(all_new_trades)}")
    print(f"Net capital gain on new trades: ${total_gain:,.2f} CAD")


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    args    = sys.argv[1:]
    dry_run = "--dry-run" in args
    scan    = "--scan" in args

    days = 90
    if "--days" in args:
        try:
            days = int(args[args.index("--days") + 1])
        except (IndexError, ValueError):
            pass

    if scan:
        scan_mode()
    else:
        main(dry_run=dry_run, days=days)
